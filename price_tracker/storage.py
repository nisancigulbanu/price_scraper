from __future__ import annotations

import csv
from pathlib import Path

from .config import settings
from .repository import latest_results

HEADERS = [
    "source_file",
    "category",
    "url",
    "domain",
    "product_name",
    "price",
    "currency",
    "raw_price",
    "quantity_grams",
    "unit_price_per_kg",
    "quantity_source",
    "quantity_confidence",
    "method",
    "confidence",
    "status",
    "error",
    "fetched_at",
    "http_status",
    "final_url",
]


def export_csv(path: Path, rows: list[object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row[header] for header in HEADERS})


def export_xlsx(path: Path, rows: list[object]) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "all_results"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row[header] for header in HEADERS])

    success = wb.create_sheet("success")
    success.append(HEADERS)
    failed = wb.create_sheet("failed")
    failed.append(HEADERS)
    needs_review = wb.create_sheet("needs_review")
    needs_review.append(HEADERS)
    for row in rows:
        target = success if row["status"] == "success" else needs_review if row["status"] == "low_confidence" else failed
        target.append([row[header] for header in HEADERS])
    wb.save(path)


def export_reports(db_path: Path | None = None, output_path: Path | None = None) -> Path:
    rows = latest_results(db_path, limit=100_000)
    output = output_path or (settings.output_dir / "prices.xlsx")
    export_csv(settings.output_dir / "prices.csv", rows)
    export_csv(settings.output_dir / "failed_urls.csv", [row for row in rows if row["status"] not in {"success", "low_confidence"}])
    export_csv(settings.output_dir / "needs_review.csv", [row for row in rows if row["status"] == "low_confidence"])
    export_xlsx(output, rows)
    return output
